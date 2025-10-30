#!/usr/bin/env python3
"""
entry_remove.py

Remove specified entries (and any linked “child” entries) from a CCDI metadata manifest.
"""

import argparse
import sys
import os
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def drop_empty(df, axis):
    """
    Remove rows (axis=0) or columns (axis=1) that are entirely NA/blank.
    """
    return df.dropna(how="all", axis=axis)

def generate_example_files():
    """
    Creates:
      - template.xlsx     (with a Dictionary sheet + blank node sheets)
      - manifest.xlsx     (sample data with linking)
      - entries.tsv       (one ID to remove)
    """
    # 1) Dictionary sheet
    dict_df = pd.DataFrame({
        'Node':     ['study',  'sample'],
        'Property': ['study_id', 'sample_type']
    })
    # add the linking property row
    extra = pd.DataFrame([{'Node': 'sample', 'Property': 'sample.study_id'}])
    dict_df = pd.concat([dict_df, extra], ignore_index=True)

    # write template.xlsx
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dictionary"
    for row in dataframe_to_rows(dict_df, index=False, header=True):
        ws.append(row)
    wb.create_sheet("study")
    wb.create_sheet("sample")
    wb.save("template.xlsx")

    # 2) manifest.xlsx
    study_df = pd.DataFrame({
        'study_id': ['S1','S2'],
        'study_name': ['Alpha','Beta']
    })
    sample_df = pd.DataFrame({
        'sample_id': ['A','B','C'],
        'sample.study_id': ['S1','S1','S2'],
        'sample_type': ['tumor','normal','tumor']
    })
    with pd.ExcelWriter("manifest.xlsx", engine="openpyxl") as writer:
        study_df.to_excel(writer, sheet_name="study", index=False)
        sample_df.to_excel(writer, sheet_name="sample", index=False)

    # 3) entries.tsv
    pd.DataFrame({'X1': ['S1']})\
      .to_csv("entries.tsv", sep="\t", index=False, header=False)

    print("Generated: template.xlsx, manifest.xlsx, entries.tsv")

def main():
    parser = argparse.ArgumentParser(
        description="Remove CCDI entries (+ children) from a manifest."
    )
    parser.add_argument("-f","--file",     help="input manifest (.xlsx,.csv,.tsv)")
    parser.add_argument("-t","--template", help="template Excel (for Dictionary sheet)")
    parser.add_argument("-e","--entry",    help="TSV of node_ids to remove")
    parser.add_argument(
      "--generate_examples",
      action="store_true",
      help="emit example files and exit"
    )

    args = parser.parse_args()
    if args.generate_examples:
        generate_example_files()
        sys.exit(0)
    if not (args.file and args.template and args.entry):
        parser.print_help()
        sys.exit(1)

    # Absolute paths & output names
    manifest_path = os.path.abspath(args.file)
    template_path = os.path.abspath(args.template)
    entry_path    = os.path.abspath(args.entry)

    base       = os.path.splitext(os.path.basename(manifest_path))[0]
    today      = datetime.date.today().strftime("%Y%m%d")
    out_xlsx   = f"{base}_EntRemove{today}.xlsx"
    log_txt    = f"{base}_EntRemove{today}_log.txt"

    # 1) Read Dictionary → list of node sheets
    dict_df = pd.read_excel(
        template_path,
        sheet_name="Dictionary",
        engine="openpyxl"
    )
    dict_df = drop_empty(dict_df, axis=0)
    node_list = dict_df['Node'].unique().tolist()

    # 2) Load each node sheet
    NA_bank = ["NA","na","N/A","n/a"]
    workbook_list = {}
    for node in node_list:
        try:
            df = pd.read_excel(
                manifest_path,
                sheet_name=node,
                dtype=str,
                na_values=NA_bank,
                keep_default_na=False,
                engine="openpyxl"
            )
        except ValueError:
            continue

        # --- FIXED here: separate df2 assignment steps ---
        df2 = drop_empty(df, axis=1)  # drop empty cols
        df2 = drop_empty(df2, axis=0)  # drop empty rows

        # require ≥1 “real” column (not just linking cols like “foo.bar_id”)
        real_cols = [c for c in df2.columns if "." not in c]
        if len(df2) > 0 and real_cols:
            workbook_list[node] = df.copy()

    if not workbook_list:
        raise RuntimeError("No valid node sheets found in manifest.")

    # 3) Read entries to remove
    entries_df = pd.read_csv(entry_path, sep="\t", header=None, names=['X1'], dtype=str)
    pending    = entries_df['X1'].dropna().tolist()

    deleted = {node: [] for node in workbook_list.keys()}

    # 4) Iterative removal + logging
    with open(log_txt, "w") as log:
        log.write("Entries to remove (and discovered children):\n")
        log.write("\n".join(pending) + "\n\n")

        while pending:
            curr = pending.pop(0)
            log.write(f"Removing: {curr}\n")
            for node, df in workbook_list.items():
                node_id_col = f"{node}_id"
                # direct hits
                if node_id_col in df.columns:
                    hits = df[df[node_id_col] == curr].index.tolist()
                    if hits:
                        deleted[node].append(curr)
                        log.write(f"  - {curr} dropped from {node}.{node_id_col}\n")
                        df.drop(index=hits, inplace=True)

                # discover child entries
                link_cols = [c for c in df.columns if "." in c and c.endswith("_id")]
                for lc in link_cols:
                    matches = df[df[lc] == curr]
                    for _, row in matches.iterrows():
                        child = row[f"{node}_id"]
                        if child not in deleted[node] and child not in pending:
                            pending.append(child)
                            log.write(f"    => discovered child {child} in {node}.{lc}\n")

        # summary
        log.write("\nSummary of deletions by sheet:\n")
        for node, items in deleted.items():
            log.write(f" {node}: {items}\n")

    # 5) Write out with openpyxl in-place
    wb = load_workbook(manifest_path)
    for node, df in workbook_list.items():
        if node in wb.sheetnames:
            ws = wb[node]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(title=node)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    # ensure at least one sheet visible
    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")
    wb.save(out_xlsx)

    print(f"\n✅ Done. Log written to {log_txt}\n   Cleaned workbook: {out_xlsx}\n")

if __name__ == "__main__":
    main()
